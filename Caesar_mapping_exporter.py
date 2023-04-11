import openpyxl

mapping = {
    "a": "L",
    "b": "M",
    "c": "N",
    "d": "O",
    "e": "P", 
    "f": "Q",
    "g": "R",
    "h": "S",
    "i": "T",
    "j": "U",
    "k": "V",
    "l": "W",
    "m": "X",
    "n": "Y",
    "o": "Z",
    "p": "A",
    "q": "B",
    "r": "C",
    "s": "D",
    "t": "E",
    "u": "F",
    "v": "G",
    "w": "H",
    "x": "I",
    "y": "J",
    "z": "K"
}

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Mapping"

for i, (letter, mapped_letter) in enumerate(mapping.items()):
    sheet.cell(row=i+1, column=1, value=letter)
    sheet.cell(row=i+1, column=2, value=mapped_letter)

wb.save("Caesar_mapping.xlsx")
